"""
Expand an uploaded .xlsx workbook into in-memory UTF-8 CSV streams so the
existing Detector / Validator / Translator pipeline can process them unchanged.
"""

import csv
import io
from datetime import date, datetime, time
from decimal import Decimal

import openpyxl

from .CSV.Detector import _detect_by_filename

# After at least one data row, stop once this many consecutive blank rows
# appear (Excel "used range" noise).
_TRAILING_EMPTY_RUN_LIMIT = 1000
# With only a header and no data rows yet, stop after this many blank rows
# so we do not walk the whole sheet.
_LEADING_EMPTY_RUN_LIMIT = 10000


def _cell_to_csv_str(val):
    if val is None:
        return ''
    if isinstance(val, bool):
        return 'True' if val else 'False'
    if isinstance(val, (int, float)):
        return str(val)
    if isinstance(val, Decimal):
        return str(val)
    if isinstance(val, datetime):
        return val.isoformat(sep=' ', timespec='seconds')
    if isinstance(val, date):
        return val.isoformat()
    if isinstance(val, time):
        return val.isoformat()
    return str(val).strip()


def _first_n_cells_empty(row, n):
    """True if the first n logical cells are blank (None, '', or whitespace-only str)."""
    for i in range(n):
        v = row[i] if row is not None and i < len(row) else None
        if v is None:
            continue
        if isinstance(v, str) and not v.strip():
            continue
        return False
    return True


def _sheet_to_csv_stringio(ws):
    """
    Stream one worksheet to CSV. Trims a long tail of blank rows caused by
    inflated worksheet dimensions in some Excel exports.

    Returns ``StringIO`` on success (possibly empty for a sheet with no rows).
    Returns ``None`` if the first row exists but is entirely blank (invalid).
    """
    it = ws.iter_rows(values_only=True)
    try:
        header_row = next(it)
    except StopIteration:
        return io.StringIO()

    hdr = [_cell_to_csv_str(c) for c in header_row]
    if not any(h.strip() for h in hdr):
        return None

    n = len(hdr)
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(hdr)

    pending_empty = []
    seen_value = False

    for row in it:
        cells = [
            _cell_to_csv_str(row[i] if row is not None and i < len(row) else None)
            for i in range(n)
        ]
        if _first_n_cells_empty(row, n):
            pending_empty.append(cells)
            limit = (
                _TRAILING_EMPTY_RUN_LIMIT if seen_value else _LEADING_EMPTY_RUN_LIMIT
            )
            if len(pending_empty) >= limit:
                break
        else:
            for pr in pending_empty:
                writer.writerow(pr)
            pending_empty = []
            seen_value = True
            writer.writerow(cells)

    buf.seek(0)
    return buf


def expand_xlsx_upload(uploaded_file):
    """
    Read a Django UploadedFile (.xlsx). Returns (list of (synthetic_filename, StringIO), errors).

    Rules:
      - At most 3 worksheets.
      - Each sheet name must classify via the same substring rules as CSV filenames
        (location, then tag, then link — case-insensitive).
      - At most one sheet per entity type.
    """
    raw = uploaded_file.read()
    bio = io.BytesIO(raw)
    try:
        # read_only=True makes iter_rows follow the workbook's stored dimensions, which
        # are often bloated to thousands of blank rows; use a normal workbook load.
        wb = openpyxl.load_workbook(bio, read_only=False, data_only=True)
    except Exception as e:
        return [], [f"Could not read Excel workbook '{uploaded_file.name}': {e}"]

    try:
        names = wb.sheetnames
        if len(names) > 3:
            return [], [
                f"Workbook '{uploaded_file.name}' has {len(names)} worksheets; at most 3 are allowed."
            ]

        seen_types = {}
        for sn in names:
            ft = _detect_by_filename(sn)
            if ft is None:
                return [], [
                    f"Worksheet '{sn}' in '{uploaded_file.name}' must include "
                    "'location', 'tag', or 'link' in its name (case-insensitive)."
                ]
            if ft in seen_types:
                return [], [
                    f"Workbook '{uploaded_file.name}' has more than one '{ft}' worksheet "
                    f"('{seen_types[ft]}' and '{sn}'). Use at most one sheet per type."
                ]
            seen_types[ft] = sn

        out = []
        for sn in names:
            ws = wb[sn]
            buf = _sheet_to_csv_stringio(ws)
            if buf is None:
                return [], [
                    f"Worksheet '{sn}' in '{uploaded_file.name}' has no usable header row."
                ]
            out.append((f"{uploaded_file.name} — {sn}", buf))

        return out, []
    finally:
        wb.close()
